from turtle import st
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Border, Side, PatternFill, GradientFill, Alignment


wb = load_workbook("Vertical_Analysis_of_Green_Delta.xlsx")
ws = wb["Sheet2"]

cell_fill = PatternFill("solid", fgColor="00C0C0C0")

for row in range(6, 28):
    if row % 2 != 0:
        for col in range(1, 8):
            char = get_column_letter(col)
            ws[char + str(row)].fill = cell_fill


for row in range(6, 28):
    corrected_20 = ""
    corrected_19 = ""
    amount_20 = str(ws["C" + str(row)].value)
    amount_19 = str(ws["F" + str(row)].value)

    for i in amount_20:
        if i == "$":
            continue
        else:
            corrected_20 += i
    ws["C" + str(row)] = corrected_20
    for i in amount_19:
        if i == "$":
            continue
        else:
            corrected_19 += i
    ws["F" + str(row)] = corrected_19



wb.save("Vertical_Analysis_of_Green_Delta.xlsx")