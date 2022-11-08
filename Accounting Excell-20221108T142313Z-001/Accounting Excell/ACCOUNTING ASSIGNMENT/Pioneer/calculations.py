from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font

wb = load_workbook("Pioneer.xlsx")
ws = wb["Financial Position"]


col_1 = 5
col_2 = 6
char_1 = get_column_letter(col_1)
char_2 = get_column_letter(col_2)

for row in range(3, 29):
    if row == 14 or row == 27 or row == 12 or row == 13 or row == 11:
        continue
    else:
        print(row)
        diff = round(float(ws["C" + str(row)].value) -  float(ws["D" + str(row)].value), 2)
        if diff < 0:
            ws[char_1 + str(row)] = "(" + "$" + str(diff * (-1)) + ")"
        else:
            ws[char_1 + str(row)] = "$" + str(diff)
        a = diff / float(ws["D" + str(row)].value)
        ws[char_2 + str(row)] = str(round(a, 2) * 100) + "%"

wb.save("Pioneer.xlsx")

