from turtle import st
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Border, Side, PatternFill, GradientFill, Alignment

wb = load_workbook("Pioneer.xlsx")
ws = wb["Financial Position"]

cell_fill = PatternFill("solid", fgColor="00C0C0C0")

for row in range(3, 25):
    if row % 2 != 0:
        for col in range(1, 7):
            char = get_column_letter(col)
            ws[char + str(row)].fill = cell_fill

wb.save("Pioneer.xlsx")


