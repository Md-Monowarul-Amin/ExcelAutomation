from turtle import st
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font

wb = load_workbook("Pioneer.xlsx")
ws = wb["Financial Position"]

for row in range(3, 23):
    corrected_string_amount = ""
    corrected_string_percent = "("
    amount = ws["E" + str(row)].value
    
    for i in amount:
        if i == "$":
            continue
        else:
            corrected_string_amount += i
    ws["E" + str(row)] = corrected_string_amount
    
    percent = str(ws["F" + str(row)].value)
    if "-" in percent:
        for i in percent:
            if i == "-":
                continue
            else:
                corrected_string_percent += i
        corrected_string_percent += ")"
        ws["F" + str(row)] = corrected_string_percent

wb.save("Pioneer.xlsx")
    


    

