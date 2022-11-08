import imp
from socketserver import DatagramRequestHandler
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font


wb = load_workbook("Pioneer.xlsx")
ws = wb["Financial Position"]
file_1 = open("Operation Result.txt", "r")

num_list = []
for i in range(30):
	string = file_1.readline(1000)
	num_list.append(string)

print(num_list)

ws.merge_cells("A14:D14")

for row in range(15,35):
	string_list = num_list[row-15]
	string_set = ""
	col = 7
	for i in reversed(range(len(string_list))):
		if string_list[i] == " ":
			if col > 1:
				char = get_column_letter(col)
				ws[char + str(row)] = string_set
				string_set = ""
				col -= 1
			else:
				string_set = string_list[i] + string_set

		else:
			string_set = string_list[i] + string_set
	ws["A" + str(row)] = string_set
		

"""for list in num_list:
	input_list = []
	string = ""
	for i in reversed(len(list)):
		if i == " ":
			input_list.append(string)
			string = ""
			continue
		else:
			string += i
	ws.append(input_list)"""

font = Font()
ws["A1"].font = Font(bold=True, color= "000000FF")
ws["A14"].font = Font(bold=True, color= "000000FF")
for col in range(1, 7):
	char = get_column_letter(col)
	ws[char + str(2)].font = Font(bold=True)

ws.delete_cols(5,7)
wb.save("Pioneer.xlsx")
