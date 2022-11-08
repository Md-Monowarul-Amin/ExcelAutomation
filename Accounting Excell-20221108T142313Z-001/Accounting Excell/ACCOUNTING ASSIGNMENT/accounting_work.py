import imp
from socketserver import DatagramRequestHandler
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font


wb = load_workbook("Green Delta.xlsx")
ws = wb["Financial Progress"]
file_1 = open("Financial Progress.txt", "r")

num_list = []
for i in range(30):
	string = file_1.readline(1000)
	num_list.append(string)

print(num_list)

ws.merge_cells("A1:C1")

for row in range(2,30):
	string_list = num_list[row-2]
	string_set = ""
	col = 6
	for i in reversed(range(len(string_list))):
		if string_list[i] == " ":
			if col > 1:
				char = get_column_letter(col)
				ws[char + str(row)] = string_set
				string_set = ""
				col -= 1
			else:
				string_set = string_list[i] + string_set

		else:
			string_set = string_list[i] + string_set
	ws["A" + str(row)] = string_set
		

"""for list in num_list:
	input_list = []
	string = ""
	for i in reversed(len(list)):
		if i == " ":
			input_list.append(string)
			string = ""
			continue
		else:
			string += i
	ws.append(input_list)"""

font = Font()
ws["A1"].font = Font(bold=True, color= "000000FF")
for col in range(1, 7):
	char = get_column_letter(col)
	ws[char + str(2)].font = Font(bold=True)

ws.delete_cols(2,3)
wb.save("Green Delta.xlsx")