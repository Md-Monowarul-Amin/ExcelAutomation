import imp
from socketserver import DatagramRequestHandler
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font

wb = load_workbook("Green Delta.xlsx")
ws_PME = wb["PME"]
ws_FP = wb["Financial Progress"]


row_FP = 3
for row in range(15, 50):
    ws_PME["A" + str(row)] = ws_FP["A" + str(row_FP)].value
    ws_PME["B" + str(row)] = ws_FP["B" + str(row_FP)].value
    ws_PME["C" + str(row)] = ws_FP["C" + str(row_FP)].value
    row_FP += 1

ws_PME.insert_rows(2)
wb.save("Green Delta.xlsx")

