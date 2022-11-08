from turtle import st
from openpyxl import load_workbook, workbook
from openpyxl.chart import BarChart, Series, Reference


def create_chart(work_sheet):
    wb = load_workbook("Ratio Analysis.xlsx")
    ws = wb[work_sheet]

    chart_1 = BarChart()
    chart_1.type = "col"
    chart_1.style = 10
    chart_1.title = ws["A1"].value + "Analysis"
    chart_1.y_axis.title = "Percentage"
    chart_1.x_axis.title = "Year"


    data = Reference(ws, min_col = 2, min_row = 2, max_row = 4, max_col = 3)
    cats = Reference(ws, min_col=1, min_row=3, max_row=4)
    chart_1.add_data(data, titles_from_data=True)
    chart_1.set_categories(cats)
    chart_1.shape = 4
    ws.add_chart(chart_1, "A6")
    print("in")
    wb.save("Ratio Analysis.xlsx")


wb_analysis = load_workbook("Ratio Analysis.xlsx")
# ws_analysis = wb_analysis["Sheet1"]

wb_pioneer = load_workbook("Pioneers_Ratio.xlsx")
ws_pioneer = wb_pioneer["Sheet1"]

wb_green_delta = load_workbook("Green_Delta_companys_Ratio.xlsx")
ws_green_delta = wb_green_delta["Sheet1"]


pioneer_dict = {}
green_delta_dict = {}
for row in range(5, 18):
    ratio = []
    key = ws_pioneer["A" + str(row)].value.upper()
    ratio.append(ws_pioneer["B" + str(row)].value)
    ratio.append(ws_pioneer["C" + str(row)].value)
    pioneer_dict[key] = ratio

for row in range(6, 16):
    ratio = []
    key = ws_green_delta["A" + str(row)].value.upper()
    ratio.append(ws_green_delta["B" + str(row)].value)
    ratio.append(ws_green_delta["C" + str(row)].value)
    green_delta_dict[key] = ratio



print(pioneer_dict )
print("\n")

print(green_delta_dict)


for ratio in pioneer_dict:
    if ratio in green_delta_dict:
        ws_analysis = wb_analysis.create_sheet()
        ws_analysis.title = ratio
        print(ws_analysis.title)
        ws_analysis["A1"] = ratio
        ws_analysis["B2"] = "Green Delta"
        ws_analysis["C2"] = "Pioneer"
        ws_analysis["A3"] = "2020"
        ws_analysis["A4"] = "2019"
        ws_analysis["B3"] = green_delta_dict[ratio][0]
        ws_analysis["B4"] = green_delta_dict[ratio][1]

        ws_analysis["C3"] = pioneer_dict[ratio][0]
        ws_analysis["C4"] = pioneer_dict[ratio][1]
        wb_analysis.save("Ratio Analysis.xlsx")

        create_chart(ws_analysis.title)
        wb_analysis.save("Ratio Analysis.xlsx")
        # print(1)

