from turtle import st
from openpyxl import load_workbook, workbook
from openpyxl.chart import BarChart, Series, Reference


wb = load_workbook("Ratio Analysis.xlsx")
ws = wb["Return On Asset"]

chart_1 = BarChart()
chart_1.type = "col"
chart_1.style = 10
chart_1.title = ws["A1"].value + "Analysis"
chart_1.y_axis.title = "Percentage"
chart_1.x_axis.title = "Year"


data = Reference(ws, min_col = 2, min_row = 2, max_row = 4, max_col = 3)
cats = Reference(ws, min_col=1, min_row=3, max_row=4)
chart_1.add_data(data, titles_from_data=True)
chart_1.set_categories(cats)
chart_1.shape = 4
ws.add_chart(chart_1, "A6")
print("in")
wb.save("Ratio Analysis.xlsx")