from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font

wb = load_workbook("Green Delta.xlsx")
ws = wb["PME"]


col_1 = 4
col_2 = 5
char_1 = get_column_letter(col_1)
char_2 = get_column_letter(col_2)

for row in range(7, 31):
    
    print(row)
    diff = round(float(ws["C" + str(row)].value) -  float(ws["B" + str(row)].value), 2)
    if diff < 0:
        ws["D" + str(row)] = "(" + str(diff * (-1)) + ")"
        a = (diff * (-1)) / float(ws["B" + str(row)].value) * 100
        ws["E" + str(row)] = "(" + str(round(a, 2)) + ")"
    else:
        ws["D" + str(row)] = str(diff)
        a = diff / float(ws["B" + str(row)].value) * 100
        ws["E" + str(row)] = str(round(a, 2))
    

wb.save("Green Delta.xlsx")
