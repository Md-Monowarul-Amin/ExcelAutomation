from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font

wb = load_workbook("Green_Delta.xlsx")
ws = wb["Income"]


for row in range(6, 19):
    
    print(row)
    diff = round(float(ws["B" + str(row)].value) -  float(ws["C" + str(row)].value), 2)
    if diff < 0:
        ws["D" + str(row)] = "(" + str(diff * (-1)) + ")"
        a = (diff * (-1)) / float(ws["B" + str(row)].value) * 100
        ws["E" + str(row)] = "(" + str(round(a, 2)) + ")"
    else:
        ws["D" + str(row)] = str(diff)
        a = diff / float(ws["B" + str(row)].value) * 100
        ws["E" + str(row)] = str(round(a, 2))



"""for sheet in wb:
    print(sheet.title)"""


wb.save("Green_Delta.xlsx")