from openpyxl import workbook, load_workbook
from openpyxl.chart import BarChart, Series, Reference


wb_analysis = load_workbook("Beximco Pharma Ratio.xlsx")
# ws_analysis = wb_analysis["Sheet1"]

# wb_pioneer = load_workbook("Pioneers_Ratio.xlsx")
ws_beximco = wb_analysis["Beximco"]
ws_square = wb_analysis["Square"]

# wb_green_delta = load_workbook("Green_Delta_companys_Ratio.xlsx")
# ws_green_delta = wb_green_delta["Sheet1"]


beximco_dict = {}
square_dict = {}
for row in range(1, 12):
    ratio = []
    key = ws_beximco["A" + str(row)].value.upper()
    ratio.append(ws_beximco["B" + str(row)].value)
    ratio.append(ws_beximco["C" + str(row)].value)
    beximco_dict[key] = ratio

for row in range(2, 12):
    ratio = []
    key = ws_square["A" + str(row)].value.upper()
    ratio.append(ws_square["B" + str(row)].value)
    ratio.append(ws_square["C" + str(row)].value)
    square_dict[key] = ratio



print(beximco_dict)
print("\n")

print(square_dict)


for ratio in beximco_dict:
    if ratio in square_dict:
        ws_analysis = wb_analysis.create_sheet()
        ws_analysis.title = ratio
        print(ws_analysis.title)
        ws_analysis["A1"] = ratio
        ws_analysis["B2"] = "Beximco"
        ws_analysis["C2"] = "Square"
        ws_analysis["A3"] = "2020"
        ws_analysis["A4"] = "2019"
        ws_analysis["B3"] = beximco_dict[ratio][0]
        ws_analysis["B4"] = beximco_dict[ratio][1]

        ws_analysis["C3"] = square_dict[ratio][0]
        ws_analysis["C4"] = square_dict[ratio][1]
        wb_analysis.save("Beximco Pharma Ratio.xlsx")

        # create_chart(ws_analysis.title)
        wb_analysis.save("Beximco Pharma Ratio.xlsx")
        # print(1)"""